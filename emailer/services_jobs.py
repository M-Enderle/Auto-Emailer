from datetime import datetime, timedelta
import asyncio
import uuid
import json
from fastapi import HTTPException

from emailer.schemas import BulkJobRequest, SendMailRequest
from emailer.services_mail import resolve_account, build_message, send_via_smtp
from emailer.utils.settings import get_settings, now_berlin
from emailer.paths import EXCEL_DIR, JOBS_DIR
import openpyxl


def is_within_work_hours(now: datetime) -> bool:
    sched = get_settings().scheduler
    if now.weekday() not in sched.workdays:
        return False
    return sched.start_hour <= now.hour < sched.end_hour


def next_allowed_time(now: datetime) -> datetime:
    sched = get_settings().scheduler
    if now.weekday() in sched.workdays and sched.start_hour <= now.hour < sched.end_hour:
        return now
    day = now
    while True:
        day = day + timedelta(days=1)
        if day.weekday() in sched.workdays:
            return day.replace(hour=sched.start_hour, minute=0, second=0, microsecond=0)


def ensure_kontaktiert_column(ws) -> int:
    max_col = ws.max_column or 1
    header_row = 1
    kontakt_col = None
    for c in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if isinstance(val, str) and val.strip().lower() == "kontaktiert":
            kontakt_col = c
            break
    if kontakt_col is None:
        kontakt_col = max_col + 1
        ws.cell(row=header_row, column=kontakt_col, value="kontaktiert")
    return kontakt_col


def load_recipients_from_excel(file_name: str) -> list[str]:
    path = EXCEL_DIR / file_name
    if not path.exists():
        raise HTTPException(status_code=404, detail="file not found")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    kontakt_col = ensure_kontaktiert_column(ws)
    # find 'email' header column
    email_col = None
    for c in range(1, (ws.max_column or 1) + 1):
        val = ws.cell(row=1, column=c).value
        if isinstance(val, str) and val.strip().lower() == "email":
            email_col = c
            break
    if email_col is None:
        raise HTTPException(status_code=400, detail="Missing required 'email' column in Excel header")
    # collect emails from the 'email' column where kontaktiert is not True
    emails: list[str] = []
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=email_col).value
        email = cell_val.strip() if isinstance(cell_val, str) else None
        if not email:
            continue
        k_val = ws.cell(row=r, column=kontakt_col).value
        if not (isinstance(k_val, bool) and k_val is True):
            emails.append(email)
    # persist header addition if newly added
    wb.save(path)
    # de-duplicate preserving order
    seen = set()
    result = []
    for e in emails:
        if e not in seen:
            seen.add(e)
            result.append(e)
    return result


def mark_contacted(file_name: str, email: str) -> None:
    path = EXCEL_DIR / file_name
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    kontakt_col = ensure_kontaktiert_column(ws)
    # find 'email' column
    email_col = None
    for c in range(1, (ws.max_column or 1) + 1):
        val = ws.cell(row=1, column=c).value
        if isinstance(val, str) and val.strip().lower() == "email":
            email_col = c
            break
    if email_col is None:
        wb.save(path)
        return
    target = (email or "").strip().lower()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=email_col).value
        if isinstance(val, str) and val.strip().lower() == target:
            ws.cell(row=r, column=kontakt_col, value=True)
            break
    wb.save(path)


def _job_path(job_id: str):
    return JOBS_DIR / f"{job_id}.json"


def _serialize_request(req: BulkJobRequest) -> dict:
    return {
        "html_body": req.html_body,
        "betreff": req.betreff,
        "from_address": req.from_address,
        "batch_size": req.batch_size,
        "interval_minutes": req.interval_minutes,
        "file_id": req.file_id,
    }


def _deserialize_request(data: dict) -> BulkJobRequest:
    return BulkJobRequest(**data)


class JobManager:
    def __init__(self):
        self.jobs: dict[str, dict] = {}

    async def add_job(self, recipients: list[str], req: BulkJobRequest) -> str:
        job_id = str(uuid.uuid4())
        self.jobs[job_id] = {
            "id": job_id,
            "recipients": recipients,
            "request": req,
            "cursor": 0,
            "sent": 0,
            "failed": 0,
            "status": "queued",
            "next_run": now_berlin().timestamp(),
            "cancelled": False,
        }
        self._save(job_id)
        asyncio.create_task(self._run_job(job_id))
        return job_id

    def list_jobs(self) -> list[dict]:
        return [self._public(j) for j in self.jobs.values()]

    def get_job(self, job_id: str) -> dict | None:
        j = self.jobs.get(job_id)
        return self._public(j) if j else None

    def _public(self, j: dict) -> dict:
        return {
            "id": j["id"],
            "total": len(j["recipients"]),
            "sent": j["sent"],
            "failed": j["failed"],
            "status": j["status"],
            "next_run": j.get("next_run"),
        }

    def _save(self, job_id: str):
        job = self.jobs.get(job_id)
        if not job:
            return
        data = {
            "id": job["id"],
            "recipients": job["recipients"],
            "request": _serialize_request(job["request"]),
            "cursor": job["cursor"],
            "sent": job["sent"],
            "failed": job["failed"],
            "status": job["status"],
            "next_run": job.get("next_run"),
            "cancelled": job.get("cancelled", False),
        }
        _job_path(job_id).write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

    def _delete_job_record(self, job_id: str) -> None:
        """Remove job from memory and delete its persisted file."""
        self.jobs.pop(job_id, None)
        try:
            _job_path(job_id).unlink(missing_ok=True)  # type: ignore[arg-type]
        except Exception:
            pass

    def load_existing(self):
        for path in JOBS_DIR.glob("*.json"):
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                job_id = data["id"]
                self.jobs[job_id] = {
                    "id": job_id,
                    "recipients": data.get("recipients", []),
                    "request": _deserialize_request(data.get("request", {})),
                    "cursor": data.get("cursor", 0),
                    "sent": data.get("sent", 0),
                    "failed": data.get("failed", 0),
                    "status": data.get("status", "queued"),
                    "next_run": data.get("next_run"),
                    "cancelled": data.get("cancelled", False),
                }
                if self.jobs[job_id]["status"] != "completed":
                    asyncio.create_task(self._run_job(job_id))
            except Exception:
                continue

    def cancel_job(self, job_id: str) -> bool:
        job = self.jobs.get(job_id)
        if not job:
            return False
        job["cancelled"] = True
        # Remove immediately from lists and disk; running loop will observe flag and exit
        self._delete_job_record(job_id)
        return True

    async def _run_job(self, job_id: str):
        while True:
            job = self.jobs.get(job_id)
            if not job:
                return
            if job.get("cancelled"):
                self._delete_job_record(job_id)
                return
            if job["cursor"] >= len(job["recipients"]):
                job["status"] = "completed"
                self._save(job_id)
                return
            now = now_berlin()
            # If a next_run is scheduled in the future, wait until then (survives restarts)
            next_run_ts = job.get("next_run")
            if isinstance(next_run_ts, (int, float)):
                now_ts = now.timestamp()
                if now_ts < next_run_ts - 0.001:
                    job["status"] = "sleeping"
                    self._save(job_id)
                    await asyncio.sleep(max(0, next_run_ts - now_ts))
                    continue
            if not is_within_work_hours(now):
                wake = next_allowed_time(now)
                job["status"] = "waiting_window"
                job["next_run"] = wake.timestamp()
                self._save(job_id)
                await asyncio.sleep(max(0, (wake - now).total_seconds()))
                continue
            req: BulkJobRequest = job["request"]
            start = job["cursor"]
            end = min(start + max(1, req.batch_size), len(job["recipients"]))
            recipients = job["recipients"][start:end]
            job["status"] = "sending"
            for r in recipients:
                if job.get("cancelled"):
                    self._delete_job_record(job_id)
                    return
                ok = await self._send_one(req, r)
                if ok:
                    job["sent"] += 1
                else:
                    job["failed"] += 1
                job["cursor"] += 1
                self._save(job_id)
            if job["cursor"] < len(job["recipients"]):
                job["status"] = "sleeping"
                wake = now + timedelta(minutes=max(0, req.interval_minutes))
                job["next_run"] = wake.timestamp()
                self._save(job_id)
                await asyncio.sleep(max(0, req.interval_minutes) * 60)
            else:
                job["status"] = "completed"
                self._save(job_id)
                return

    async def _send_one(self, req: BulkJobRequest, recipient: str) -> bool:
        payload = SendMailRequest(
            html_body=req.html_body,
            betreff=req.betreff,
            recipient=recipient,  # type: ignore[arg-type]
            from_address=req.from_address,
        )
        try:
            account = resolve_account(payload.from_address)
            msg = build_message(account, payload)
            await asyncio.to_thread(send_via_smtp, account, msg)
            mark_contacted(req.file_id, recipient)
            return True
        except Exception:
            return False


